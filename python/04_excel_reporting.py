# 04_excel_report.py
# Layer 2 — Automated Excel Exception Report
# Reads transactions.csv, scores each exception using Layer 1 logic,
# and generates a formatted audit-quality Excel report.

import csv
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date
from risk_scoring import score_exception

# ---------------------------------------------------------------------------
# Color constants (hex codes, no # prefix for openpyxl)
# ---------------------------------------------------------------------------
HIGH_COLOR = "FF9999"   # soft red
MED_COLOR  = "FDFD96"   # yellow
LOW_COLOR  = "C1E1C1"   # soft green
HEADER_COLOR = "2F4F8F" # dark navy for column headers
TITLE_COLOR  = "1C1C3A" # near black for title row

# ---------------------------------------------------------------------------
# Which exception types to flag (mirrors your SQL control logic)
# ---------------------------------------------------------------------------
EXCEPTION_TYPES = {
    "approval_breach":  "Approval Limit Breach",
    "duplicate":        "Duplicate Invoice Payment",
    "split":            "Potential Split Payment",
}


# ---------------------------------------------------------------------------
# Step 1 — Load transactions and identify exceptions
# Mirrors the three SQL control tests from 02_controls.sql
# ---------------------------------------------------------------------------
def load_exceptions(filepath):
    exceptions = []

    with open(filepath, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)

    # Control 1 — Approval Limit Breaches
    for row in rows:
        amount = float(row["amount"])
        limit  = float(row["approval_limit"])
        if amount > limit:
            exceptions.append({
                "transaction_id":   row["transaction_id"],
                "transaction_date": row["transaction_date"],
                "exception_type":   EXCEPTION_TYPES["approval_breach"],
                "department":       row["department"],
                "vendor_name":      row["vendor_name"],
                "vendor_country":   row["vendor_country"],
                "amount":           amount,
                "approval_limit":   limit,
            })

    # Control 2 — Duplicate Invoice Payments
    # Same vendor + same invoice_id appearing more than once
    seen_invoices = {}
    for row in rows:
        key = (row["vendor_id"], row["invoice_id"])
        seen_invoices.setdefault(key, []).append(row)

    for key, dupes in seen_invoices.items():
        if len(dupes) > 1:
            for row in dupes:
                exceptions.append({
                    "transaction_id":   row["transaction_id"],
                    "transaction_date": row["transaction_date"],
                    "exception_type":   EXCEPTION_TYPES["duplicate"],
                    "department":       row["department"],
                    "vendor_name":      row["vendor_name"],
                    "vendor_country":   row["vendor_country"],
                    "amount":           float(row["amount"]),
                    "approval_limit":   float(row["approval_limit"]),
                })

    # Control 3 — Potential Split Payments
    # Same vendor + same date, total amount > 10,000
    vendor_day = {}
    for row in rows:
        key = (row["vendor_id"], row["transaction_date"])
        vendor_day.setdefault(key, []).append(row)

    for key, group in vendor_day.items():
        if len(group) > 1 and sum(float(r["amount"]) for r in group) > 10000:
            for row in group:
                exceptions.append({
                    "transaction_id":   row["transaction_id"],
                    "transaction_date": row["transaction_date"],
                    "exception_type":   EXCEPTION_TYPES["split"],
                    "department":       row["department"],
                    "vendor_name":      row["vendor_name"],
                    "vendor_country":   row["vendor_country"],
                    "amount":           float(row["amount"]),
                    "approval_limit":   float(row["approval_limit"]),
                })

    return exceptions


# ---------------------------------------------------------------------------
# Step 2 — Score every exception using Layer 1
# ---------------------------------------------------------------------------
def apply_scores(exceptions):
    for row in exceptions:
        row["severity"] = score_exception(
            exception_type  = row["exception_type"],
            amount          = row["amount"],
            approval_limit  = row["approval_limit"],
            vendor_country  = row["vendor_country"],
        )
    return exceptions


# ---------------------------------------------------------------------------
# Step 3 — Sort by amount descending (highest dollar exposure first)
# ---------------------------------------------------------------------------
def sort_by_amount(exceptions):
    return sorted(exceptions, key=lambda row: row["amount"], reverse=True)


# ---------------------------------------------------------------------------
# Step 4 — Build the Excel report
# ---------------------------------------------------------------------------
def build_report(exceptions, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Exception Report"

    # -- Severity fill colors --
    fills = {
        "High":   PatternFill("solid", fgColor=HIGH_COLOR),
        "Medium": PatternFill("solid", fgColor=MED_COLOR),
        "Low":    PatternFill("solid", fgColor=LOW_COLOR),
    }

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # -----------------------------------------------------------------------
    # Step 4a — Header section (title, date, summary)
    # -----------------------------------------------------------------------
    total       = len(exceptions)
    total_exp   = sum(r["amount"] for r in exceptions)
    high_count  = sum(1 for r in exceptions if r["severity"] == "High")
    med_count   = sum(1 for r in exceptions if r["severity"] == "Medium")
    low_count   = sum(1 for r in exceptions if r["severity"] == "Low")

    ws.merge_cells("A1:G1")
    ws["A1"] = "Transaction Monitoring — Control Exception Report"
    ws["A1"].font      = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor=TITLE_COLOR)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:G2")
    ws["A2"] = f"Generated: {date.today().strftime('%B %d, %Y')}"
    ws["A2"].font      = Font(italic=True, size=11, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = (
        f"Total Exceptions: {total}     |     "
        f"Total Exposure: ${total_exp:,.2f}     |     "
        f"High: {high_count}     Medium: {med_count}     Low: {low_count}"
    )
    ws["A3"].font      = Font(bold=True, size=11)
    ws["A3"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[3].height = 20

    ws.append([])  # blank row 4 for breathing room

    # -----------------------------------------------------------------------
    # Step 4b — Column headers
    # -----------------------------------------------------------------------
    headers = [
        "Transaction ID",
        "Date",
        "Exception Type",
        "Department",
        "Vendor Country",
        "Amount ($)",
        "Severity",
    ]
    ws.append(headers)

    header_row = 5
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font      = Font(bold=True, color="FFFFFF", size=11)
        cell.fill      = PatternFill("solid", fgColor=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = thin_border
    ws.row_dimensions[header_row].height = 20

    # -----------------------------------------------------------------------
    # Step 4c — Data rows, color coded by severity
    # -----------------------------------------------------------------------
    for exc in exceptions:
        row_data = [
            int(exc["transaction_id"]),
            exc["transaction_date"],
            exc["exception_type"],
            exc["department"],
            exc["vendor_country"],
            round(exc["amount"], 2),
            exc["severity"],
        ]
        ws.append(row_data)

        current_row = ws.max_row
        fill = fills[exc["severity"]]

        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal="center")

        # Format the amount column as currency
        ws.cell(row=current_row, column=6).number_format = '#,##0.00'

    # -----------------------------------------------------------------------
    # Step 4d — Column widths
    # -----------------------------------------------------------------------
    col_widths = [16, 14, 28, 14, 16, 14, 12]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    # -----------------------------------------------------------------------
    # Step 4e — Freeze panes so header stays visible when scrolling
    # -----------------------------------------------------------------------
    ws.freeze_panes = "A6"

    # -----------------------------------------------------------------------
    # Step 4f — Save
    # -----------------------------------------------------------------------
    wb.save(output_path)
    print(f"\nReport saved → {output_path}")
    print(f"  {total} exceptions  |  ${total_exp:,.2f} total exposure")
    print(f"  High: {high_count}  |  Medium: {med_count}  |  Low: {low_count}")


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    exceptions = load_exceptions("../data/transactions.csv")
    exceptions = apply_scores(exceptions)
    exceptions = sort_by_amount(exceptions)
    build_report(exceptions, "../data/exception_report.xlsx")
